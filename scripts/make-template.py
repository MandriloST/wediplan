#!/usr/bin/env python3
"""Generira data/vendors-template.xlsx — radni predložak za istraživanje pružatelja."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

CATEGORIES = [
    "Restorani i sale", "Konobe i prostori za proslavu", "Najam kuće za proslavu",
    "Najam šatora i opreme", "Catering", "Torte i kolači", "Foto i Video",
    "Audio, foto kabine i selfie mirror", "Glazba — bendovi", "DJ", "Glazba za crkvu",
    "Harmonikaši", "Cvijeće, dekoracije i baloni", "Vjenčanice i dodaci",
    "Muška odijela i dodaci", "Nakit i prstenje", "Frizerski saloni",
    "Šminka, nokti i trepavice",
    "Najam auta, limuzina i oldtimera", "Prijevoz i transferi", "Organizatori vjenčanja",
    "Pozivnice, zahvalnice i popis gostiju", "Rasvjeta, razglas i efekti",
    "Čuvanje i animacija djece", "Škole plesa", "Bračna putovanja",
    "Dodaci za djevojačke i momačke", "Reveri, narukvice, podvezice i dodaci",
]
REGIONS = ["Istra", "Kvarner", "Dalmacija", "Zagreb i okolica", "Slavonija"]
STATUSES = ["Istraženo", "Kontaktirano", "Dozvola dobivena", "Objavljeno"]

HEADERS = [
    ("naziv*", 26, "Puni naziv pružatelja"),
    ("kategorija*", 24, "Odaberi iz padajućeg izbornika"),
    ("regija*", 16, "Odaberi iz padajućeg izbornika"),
    ("grad*", 14, None),
    ("koordinate*", 20, 'Google Maps → desni klik na lokaciju → klik na koordinate (kopira "43.5081, 16.4402") → zalijepi ovdje'),
    ("nacin_cijene*", 16, "od (paušal) = jedna cijena 'od X €'; po osobi (raspon) = X–Y €/os. (sale, catering)"),
    ("cijena_od*", 11, "Samo broj u €, bez oznake valute"),
    ("cijena_do", 11, "Samo za 'po osobi (raspon)'"),
    ("ocjena", 9, "0–5, npr. 4.8 — prenesena ocjena (Google i sl.). Ostavi prazno ako nema."),
    ("broj_recenzija", 13, None),
    ("izvor_ocjene", 16, "OBAVEZNO ako je ocjena upisana, npr. 'Google recenzije'"),
    ("provjereno", 12, "DA = kontaktirali smo ih i potvrdili podatke"),
    ("kalendar_uzivo", 14, "Zasad NE za sve — DA tek kad pružatelj dobije CRM"),
    ("stil", 22, "Tagovi odvojeni zarezom, npr: uz more, terasa"),
    ("o_pruzatelju", 46, "Što pružatelj kaže o sebi (2–4 rečenice). Prazno = automatski tekst po kategoriji."),
    ("usluge", 34, "Odvojeno zarezom. Prazno = automatski popis po kategoriji."),
    ("web", 22, "INTERNO — ne prikazuje se na stranici"),
    ("telefon", 15, "INTERNO — ne prikazuje se na stranici"),
    ("email", 22, "INTERNO — ne prikazuje se na stranici"),
    ("status", 18, "Za tvoje praćenje — ne utječe na import"),
    ("napomena", 30, "Interno"),
]

REVIEW_HEADERS = [
    ("naziv_pruzatelja*", 26, "Mora točno odgovarati nazivu s lista Pružatelji"),
    ("autor*", 18, "npr. 'Marija i Ivan' ili 'Petra K.'"),
    ("ocjena*", 9, "1–5"),
    ("tekst*", 60, None),
    ("izvor*", 18, "npr. 'Google recenzije', 'Facebook'"),
    ("godina*", 9, None),
]

INK, PRIMARY, TINT, WARN = "21252C", "2470CC", "E9F1FB", "FFF3CD"
head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
head_fill = PatternFill("solid", start_color=PRIMARY)
base_font = Font(name="Arial", size=10)
thin = Border(bottom=Side(style="thin", color="D4CDBD"))

wb = Workbook()

# ---------------- Upute ----------------
up = wb.active
up.title = "Upute"
up.column_dimensions["A"].width = 110
lines = [
    ("WEDIPLAN — predložak za unos pružatelja", True),
    ("", False),
    ("KAKO RADITI:", True),
    ("1. Popunjavaj list 'Pružatelji' — stupci s * su obavezni. Ostalo može ostati prazno.", False),
    ("2. Koordinate: Google Maps → desni klik na lokaciju → klikni koordinate (automatski se kopiraju) → zalijepi u ćeliju.", False),
    ("3. Prenesene recenzije (tekstualne) upisuj u list 'Recenzije' — poveži ih točnim nazivom pružatelja.", False),
    ("4. Kad želiš podatke na stranici, spremi ovu datoteku i pokreni:  npm run import:vendors -- putanja/do/datoteke.xlsx", False),
    ("   Skripta provjerava svaki redak (koordinate, kategorije, cijene…) i ispisuje greške s brojem retka.", False),
    ("5. Ako je sve u redu, generira data/vendors.json i data/profiles.json → git commit → podaci su na stranici.", False),
    ("", False),
    ("VAŽNO — DOZVOLE:", True),
    ("• Za objavu prenesenih recenzija, fotografija i ocjena zatraži suglasnost pružatelja (to je ujedno prvi kontakt za partnerstvo).", False),
    ("• Kolone web/telefon/email su samo za tvoju evidenciju — import ih NE stavlja na stranicu (Direktan kontakt: zasad ne).", False),
    ("• Redci čiji naziv počinje s 'PRIMJER' se preskaču pri importu — slobodno ih ostavi ili obriši.", False),
    ("", False),
    ("SAVJET ZA POČETAK:", True),
    ("• Kreni s 1 regijom i 2–3 kategorije (npr. Dalmacija: Restorani i sale + Foto i Video) da provjeriš cijeli tok, pa širi.", False),
]
for i, (txt, bold) in enumerate(lines, 1):
    c = up.cell(row=i, column=1, value=txt)
    c.font = Font(name="Arial", size=11 if bold else 10, bold=bold, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ---------------- Šifrarnici (skriveni) ----------------
sif = wb.create_sheet("Sifrarnici")
for i, cat in enumerate(CATEGORIES, 1):
    sif.cell(row=i, column=1, value=cat)
for i, r in enumerate(REGIONS, 1):
    sif.cell(row=i, column=2, value=r)
for i, s in enumerate(STATUSES, 1):
    sif.cell(row=i, column=3, value=s)
sif.sheet_state = "hidden"

# ---------------- Pružatelji ----------------
ws = wb.create_sheet("Pružatelji")
for col, (name, width, note) in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=name)
    c.font, c.fill = head_font, head_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = width
    if note:
        from openpyxl.comments import Comment
        c.comment = Comment(note, "Wediplan", height=120, width=280)
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 28

MAXR = 500
dv_cat = DataValidation(type="list", formula1=f"Sifrarnici!$A$1:$A${len(CATEGORIES)}", allow_blank=True, showErrorMessage=True)
dv_reg = DataValidation(type="list", formula1=f"Sifrarnici!$B$1:$B${len(REGIONS)}", allow_blank=True, showErrorMessage=True)
dv_stat = DataValidation(type="list", formula1=f"Sifrarnici!$C$1:$C${len(STATUSES)}", allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"DA,NE"', allow_blank=True)
dv_price = DataValidation(type="list", formula1='"od (paušal),po osobi (raspon)"', allow_blank=True)
for dv, col in [(dv_cat, "B"), (dv_reg, "C"), (dv_price, "F"), (dv_yn, "L"), (dv_yn, "M"), (dv_stat, "T")]:
    pass
# DataValidation objekti se ne smiju dijeliti za više raspona istog tipa različitih stupaca — kreiraj zasebno
dv_yn2 = DataValidation(type="list", formula1='"DA,NE"', allow_blank=True)
ws.add_data_validation(dv_cat); dv_cat.add(f"B2:B{MAXR}")
ws.add_data_validation(dv_reg); dv_reg.add(f"C2:C{MAXR}")
ws.add_data_validation(dv_price); dv_price.add(f"F2:F{MAXR}")
ws.add_data_validation(dv_yn); dv_yn.add(f"L2:L{MAXR}")
ws.add_data_validation(dv_yn2); dv_yn2.add(f"M2:M{MAXR}")
ws.add_data_validation(dv_stat); dv_stat.add(f"T2:T{MAXR}")

examples = [
    ["PRIMJER — Villa Dalmacija", "Restorani i sale", "Dalmacija", "Split", "43.5147, 16.4102",
     "po osobi (raspon)", 65, 95, 4.8, 57, "Google recenzije", "DA", "NE", "uz more, terasa",
     "Terasa uz more za do 220 gostiju, vlastita kuhinja i parking. Cijena po osobi uključuje meni od 5 slijedova.",
     "Meni po osobi, Osoblje, Osnovna dekoracija, Parking", "villa-dalmacija.hr", "021/555-123",
     "info@villa-dalmacija.hr", "Dozvola dobivena", "primjer — obriši ili ostavi (preskače se)"],
    ["PRIMJER — Foto studio Anić", "Foto i Video", "Dalmacija", "Split", "43.5081, 16.4402",
     "od (paušal)", 850, None, 4.8, 31, "Google recenzije", "DA", "NE", "boho, film",
     "Vjenčanja fotografiramo od 2014. — reportažno, s naglaskom na svjetlo i emociju.",
     "", "fotostudio-anic.hr", "", "", "Kontaktirano", "primjer — preskače se pri importu"],
]
for r, row in enumerate(examples, 2):
    for col, val in enumerate(row, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10, italic=True, color="6B6558")
        c.fill = PatternFill("solid", start_color="FBFAF7")
        c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical="top")

# ---------------- Recenzije ----------------
rw = wb.create_sheet("Recenzije")
for col, (name, width, note) in enumerate(REVIEW_HEADERS, 1):
    c = rw.cell(row=1, column=col, value=name)
    c.font, c.fill = head_font, head_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    rw.column_dimensions[get_column_letter(col)].width = width
    if note:
        from openpyxl.comments import Comment
        c.comment = Comment(note, "Wediplan", height=90, width=260)
rw.freeze_panes = "A2"
rev_examples = [
    ["PRIMJER — Villa Dalmacija", "Ana i Marko", 5,
     "Terasa uz more je spektakularna, hrana odlična, osoblje na razini.", "Google recenzije", 2025],
]
for r, row in enumerate(rev_examples, 2):
    for col, val in enumerate(row, 1):
        c = rw.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10, italic=True, color="6B6558")
        c.fill = PatternFill("solid", start_color="FBFAF7")

wb.save("data/vendors-template.xlsx")
print("template saved")
